"""テスト用のExcelファイルを作成するスクリプト"""

import openpyxl
from pathlib import Path

def create_test_excel_v1():
    """バージョン1のテストExcelを作成"""
    wb = openpyxl.Workbook()

    # シート1: 基本設定
    ws1 = wb.active
    ws1.title = "基本設定"
    ws1.append(["設定項目", "値"])
    ws1.append(["ホスト名", "server1"])
    ws1.append(["ポート", "8080"])
    ws1.append(["タイムアウト", "30"])

    # シート2: 監視設定
    ws2 = wb.create_sheet("監視設定")
    ws2.append(["監視ID", "項目名", "閾値"])
    ws2.append(["MON_001", "CPU使用率", "80"])
    ws2.append(["MON_002", "メモリ使用率", "90"])
    ws2.append(["MON_003", "ディスク使用率", "95"])

    # シート3: ジョブ定義
    ws3 = wb.create_sheet("ジョブ定義")
    ws3.append(["ジョブID", "ジョブ名", "実行時刻"])
    ws3.append(["JOB_001", "バックアップ", "02:00"])
    ws3.append(["JOB_002", "ログ削除", "03:00"])

    wb.save("test_config_v1.xlsx")
    print("Created: test_config_v1.xlsx")

def create_test_excel_v2():
    """バージョン2のテストExcelを作成（変更を加えた版）"""
    wb = openpyxl.Workbook()

    # シート1: 基本設定（ポート番号を変更）
    ws1 = wb.active
    ws1.title = "基本設定"
    ws1.append(["設定項目", "値"])
    ws1.append(["ホスト名", "server1"])
    ws1.append(["ポート", "8081"])  # 変更: 8080 -> 8081
    ws1.append(["タイムアウト", "30"])

    # シート2: 監視設定（行を追加・削除・変更）
    ws2 = wb.create_sheet("監視設定")
    ws2.append(["監視ID", "項目名", "閾値"])
    ws2.append(["MON_001", "CPU使用率", "80"])
    ws2.append(["MON_004", "ネットワーク", "100"])  # 追加: 新しい監視項目
    ws2.append(["MON_002", "メモリ使用率", "85"])  # 変更: 90 -> 85
    # MON_003は削除

    # シート3: ジョブ定義（変更なし）
    ws3 = wb.create_sheet("ジョブ定義")
    ws3.append(["ジョブID", "ジョブ名", "実行時刻"])
    ws3.append(["JOB_001", "バックアップ", "02:00"])
    ws3.append(["JOB_002", "ログ削除", "03:00"])

    # シート4: 新規シート追加
    ws4 = wb.create_sheet("アラート設定")
    ws4.append(["アラートID", "通知先"])
    ws4.append(["ALT_001", "admin@example.com"])

    wb.save("test_config_v2.xlsx")
    print("Created: test_config_v2.xlsx")

if __name__ == "__main__":
    create_test_excel_v1()
    create_test_excel_v2()
    print("\n変更内容:")
    print("- 基本設定: ポート 8080 -> 8081")
    print("- 監視設定: MON_003削除、MON_004追加、MON_002の閾値変更")
    print("- アラート設定シート追加")
