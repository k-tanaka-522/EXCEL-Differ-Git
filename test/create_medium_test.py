"""中規模テスト用のExcelファイルを作成するスクリプト

メモリ使用量を抑えた100×100データで10シート、100行程度の変更をテストします。
"""

import openpyxl
import random
import string

def random_string(length=10):
    """ランダムな文字列を生成"""
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))

def random_number():
    """ランダムな数値を生成"""
    return random.randint(1, 10000)

def random_formula(row):
    """ランダムな数式を生成"""
    formulas = [
        f"=SUM(A{row}:C{row})",
        f"=AVERAGE(A{row}:E{row})",
        f"=IF(A{row}>50, \"High\", \"Low\")",
        f"=A{row}*B{row}",
    ]
    return random.choice(formulas)

def create_medium_excel_v1():
    """バージョン1: 100×100データ、10シート"""
    print("Creating medium test file v1...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_num in range(1, 11):
        print(f"  Creating sheet {sheet_num}/10...")
        ws = wb.create_sheet(f"Data{sheet_num:02d}")

        # ヘッダー
        header = ["ID"] + [f"Col{i}" for i in range(1, 100)]
        ws.append(header)

        # データ行（100行）
        for row_num in range(1, 101):
            row_data = [f"ID_{sheet_num:02d}_{row_num:03d}"]
            for col in range(1, 100):
                if col <= 10:
                    row_data.append(random_string(6))
                elif col == 20:
                    if row_num % 5 == 0:
                        row_data.append(random_formula(row_num + 1))
                    else:
                        row_data.append(random_number())
                else:
                    row_data.append(random_number())
            ws.append(row_data)

    filename = "medium_test_v1.xlsx"
    wb.save(filename)
    print(f"Created: {filename}")
    return filename

def create_medium_excel_v2():
    """バージョン2: v1をベースに100行程度の変更を加える"""
    print("\nCreating medium test file v2 (with changes)...")
    wb = openpyxl.load_workbook("medium_test_v1.xlsx")

    changes_log = []

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  Modifying {sheet_name}...")
        ws = wb[sheet_name]

        # 行削除: 10行
        for _ in range(10):
            delete_row = random.randint(10, 90)
            ws.delete_rows(delete_row)
        changes_log.append(f"{sheet_name}: Deleted 10 rows")

        # 行挿入: 15行
        for _ in range(15):
            insert_pos = random.randint(10, 50)
            new_row = [f"NEW_{sheet_idx:02d}_{random.randint(1, 999):03d}"]
            for col in range(1, 100):
                new_row.append(random_string(6) + "_N" if col <= 10 else random_number())

            ws.insert_rows(insert_pos)
            for col_idx, value in enumerate(new_row, 1):
                ws.cell(row=insert_pos, column=col_idx, value=value)
        changes_log.append(f"{sheet_name}: Inserted 15 rows")

        # セル更新: 75箇所
        for _ in range(75):
            row_num = random.randint(2, min(ws.max_row, 100))
            col_num = random.randint(2, 20)
            new_value = random_string(6) + "_M" if col_num <= 10 else random_number()
            ws.cell(row=row_num, column=col_num, value=new_value)
        changes_log.append(f"{sheet_name}: Modified 75 cells")

    filename = "medium_test_v2.xlsx"
    wb.save(filename)
    print(f"Created: {filename}")

    print(f"\n=== Change Summary ===")
    for log in changes_log:
        print(log)
    print(f"Per sheet: 10 deletions + 15 insertions + 75 modifications = 100 changes")

    return filename

if __name__ == "__main__":
    print("=== Medium Excel Test File Generator ===")
    print("Specifications:")
    print("- Size: 100 rows × 100 columns")
    print("- Sheets: 10")
    print("- Changes: ~100 per sheet (10 deletions + 15 insertions + 75 updates)")
    print()

    v1_file = create_medium_excel_v1()
    v2_file = create_medium_excel_v2()

    print("\n=== Test files created successfully! ===")
    print(f"Old version: {v1_file}")
    print(f"New version: {v2_file}")
    print("\nYou can now test with:")
    print(f"  py -m excel_differ.cli --old {v1_file} --new {v2_file} {v1_file}")
