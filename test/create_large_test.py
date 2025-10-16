"""大規模テスト用のExcelファイルを作成するスクリプト

1000×1000のデータで10シート、100行程度の変更をテストします。
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
import random
import string

def random_string(length=10):
    """ランダムな文字列を生成"""
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))

def random_number():
    """ランダムな数値を生成"""
    return random.randint(1, 100000)

def random_formula(row):
    """ランダムな数式を生成"""
    formulas = [
        f"=SUM(A{row}:C{row})",
        f"=AVERAGE(A{row}:E{row})",
        f"=IF(A{row}>50, \"High\", \"Low\")",
        f"=A{row}*B{row}",
        f"=MAX(A{row}:D{row})",
        f"=MIN(A{row}:D{row})",
        f"=ROUND(A{row}/100, 2)",
    ]
    return random.choice(formulas)

def create_large_excel_v1():
    """バージョン1: 1000×1000データ、10シート"""
    print("Creating large test file v1...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # 10シート作成
    for sheet_num in range(1, 11):
        print(f"  Creating sheet {sheet_num}/10...")
        ws = wb.create_sheet(f"Sheet{sheet_num:02d}")

        # ヘッダー行（1000列）
        header = []
        for col in range(1000):
            if col == 0:
                header.append("ID")
            elif col < 10:
                header.append(f"Data{col}")
            elif col < 20:
                header.append(f"Value{col}")
            elif col == 20:
                header.append("Formula")
            else:
                header.append(f"Col{col}")
        ws.append(header)

        # データ行（1000行）
        for row_num in range(1, 1001):
            if row_num % 100 == 0:
                print(f"    Row {row_num}/1000...")

            row_data = []
            for col in range(1000):
                if col == 0:
                    # ID列
                    row_data.append(f"ID_{sheet_num:02d}_{row_num:04d}")
                elif col < 10:
                    # 文字列データ
                    row_data.append(random_string(8))
                elif col < 20:
                    # 数値データ
                    row_data.append(random_number())
                elif col == 20:
                    # 数式（一部の行のみ）
                    if row_num % 10 == 0:
                        row_data.append(random_formula(row_num + 1))
                    else:
                        row_data.append(random_number())
                else:
                    # その他のデータ
                    if random.random() < 0.7:
                        row_data.append(random_number())
                    else:
                        row_data.append(random_string(5))

            ws.append(row_data)

    filename = "large_test_v1.xlsx"
    print(f"Saving {filename}...")
    wb.save(filename)
    print(f"Created: {filename}")
    return filename

def create_large_excel_v2():
    """バージョン2: v1をベースに100行程度の変更を加える"""
    print("\nCreating large test file v2 (with changes)...")

    # v1を読み込み
    print("Loading v1...")
    wb = openpyxl.load_workbook("large_test_v1.xlsx")

    changes_log = []
    total_changes = 0

    # 各シートに変更を加える
    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  Modifying {sheet_name}...")
        ws = wb[sheet_name]
        sheet_changes = 0

        # 変更の種類と数
        # 1. 行削除: 10行
        deleted_rows = random.sample(range(100, 900), 10)
        deleted_rows.sort(reverse=True)  # 後ろから削除
        for row_num in deleted_rows:
            ws.delete_rows(row_num)
            sheet_changes += 1
        changes_log.append(f"{sheet_name}: Deleted 10 rows")

        # 2. 行挿入: 15行
        for _ in range(15):
            insert_pos = random.randint(50, 500)
            new_row = []
            for col in range(1000):
                if col == 0:
                    new_row.append(f"NEW_{sheet_idx:02d}_{random.randint(1, 9999):04d}")
                elif col < 10:
                    new_row.append(random_string(8) + "_NEW")
                elif col < 20:
                    new_row.append(random_number())
                else:
                    new_row.append(random_number() if random.random() < 0.7 else random_string(5))

            ws.insert_rows(insert_pos)
            for col_idx, value in enumerate(new_row, 1):
                ws.cell(row=insert_pos, column=col_idx, value=value)
            sheet_changes += 1
        changes_log.append(f"{sheet_name}: Inserted 15 rows")

        # 3. セル更新: 各シート約75箇所
        for _ in range(75):
            row_num = random.randint(2, min(ws.max_row, 1000))
            col_num = random.randint(2, min(20, ws.max_column))  # 主要な列のみ

            old_value = ws.cell(row=row_num, column=col_num).value
            if col_num < 10:
                # 文字列列
                new_value = random_string(8) + "_MOD"
            else:
                # 数値列
                new_value = random_number()

            ws.cell(row=row_num, column=col_num, value=new_value)
            sheet_changes += 1
        changes_log.append(f"{sheet_name}: Modified 75 cells")

        total_changes += sheet_changes
        print(f"    Changes: {sheet_changes}")

    filename = "large_test_v2.xlsx"
    print(f"Saving {filename}...")
    wb.save(filename)
    print(f"Created: {filename}")

    print(f"\n=== Change Summary ===")
    for log in changes_log:
        print(log)
    print(f"Total changes: {total_changes}")
    print(f"Per sheet average: 10 deletions + 15 insertions + 75 modifications = 100 changes")

    return filename

if __name__ == "__main__":
    print("=== Large Excel Test File Generator ===")
    print("Specifications:")
    print("- Size: 1000 rows × 1000 columns")
    print("- Sheets: 10")
    print("- Changes: ~100 per sheet (10 deletions + 15 insertions + 75 updates)")
    print("- Formulas: Included in column 21")
    print()

    v1_file = create_large_excel_v1()
    v2_file = create_large_excel_v2()

    print("\n=== Test files created successfully! ===")
    print(f"Old version: {v1_file}")
    print(f"New version: {v2_file}")
    print("\nYou can now test with:")
    print(f"  py -m excel_differ.cli --old {v1_file} --new {v2_file} {v1_file}")
