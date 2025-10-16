"""Excel file reading functionality."""

from typing import List, Dict, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelRow:
    """Represents a single row in an Excel sheet."""

    def __init__(self, row_number: int, cells: List[Any]):
        self.row_number = row_number
        self.cells = cells

    def to_string(self) -> str:
        """Convert row to string representation for comparison."""
        return "|".join(str(cell) if cell is not None else "" for cell in self.cells)

    def __eq__(self, other) -> bool:
        """Check if two rows have the same content."""
        if not isinstance(other, ExcelRow):
            return False
        return self.cells == other.cells

    def __hash__(self) -> int:
        """Hash based on cell content."""
        return hash(tuple(str(cell) for cell in self.cells))

    def __repr__(self) -> str:
        return f"Row({self.row_number}: {self.to_string()})"


class ExcelSheet:
    """Represents a single sheet in an Excel workbook."""

    def __init__(self, name: str, rows: List[ExcelRow]):
        self.name = name
        self.rows = rows

    def __repr__(self) -> str:
        return f"Sheet({self.name}, {len(self.rows)} rows)"


class ExcelWorkbook:
    """Represents an Excel workbook."""

    def __init__(self, filepath: Path, sheets: Dict[str, ExcelSheet]):
        self.filepath = filepath
        self.sheets = sheets

    def __repr__(self) -> str:
        return f"Workbook({self.filepath.name}, {len(self.sheets)} sheets)"


def read_excel_file(filepath: Path) -> ExcelWorkbook:
    """
    Read an Excel file and return a structured representation.

    Args:
        filepath: Path to the Excel file

    Returns:
        ExcelWorkbook object containing all sheets and rows
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = {}

    for sheet_name in wb.sheetnames:
        ws: Worksheet = wb[sheet_name]
        rows = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Convert row tuple to list and store
            cells = list(row)
            excel_row = ExcelRow(row_idx, cells)
            rows.append(excel_row)

        sheets[sheet_name] = ExcelSheet(sheet_name, rows)

    wb.close()
    return ExcelWorkbook(filepath, sheets)


def read_excel_from_bytes(file_bytes: bytes, filename: str = "temp.xlsx") -> ExcelWorkbook:
    """
    Read an Excel file from bytes (useful for Git blob reading).

    Args:
        file_bytes: Excel file content as bytes
        filename: Virtual filename for reference

    Returns:
        ExcelWorkbook object containing all sheets and rows
    """
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    sheets = {}

    for sheet_name in wb.sheetnames:
        ws: Worksheet = wb[sheet_name]
        rows = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = list(row)
            excel_row = ExcelRow(row_idx, cells)
            rows.append(excel_row)

        sheets[sheet_name] = ExcelSheet(sheet_name, rows)

    wb.close()
    return ExcelWorkbook(Path(filename), sheets)
